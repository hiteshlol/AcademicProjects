import openpyxl
from openpyxl import Workbook
from datetime import datetime

"""Hotel Management System Module: 

This module implements a simple hotel management system that tracks room availability, assigns rooms to guests, and displays information about current guests.

Classes:
- HotelManagementSystem: Represents the hotel management system, including methods for loading data, checking room availability, assigning rooms, and displaying guest information.

Methods:
- load_data(): Loads data from an existing Excel file or creates a new one if not found.
- check_availability(): Checks if there are available rooms in the hotel.
- assign_room(guest_name, checkout_date): Assigns a room to a guest based on availability and input information.
- display_guests(): Displays information about current guests.

"""

class HotelManagementSystem:
    def __init__(self, total_rooms=10):
        # Initialize the HotelManagementSystem with a specified or default number of total rooms
        self.total_rooms = total_rooms
        # Load data from the existing Excel file or create a new one if not found
        self.load_data()

    def load_data(self):
        try:
            # Try loading the existing workbook from the file
            self.workbook = openpyxl.load_workbook('hotel_data.xlsx')
            self.sheet = self.workbook.active
        except FileNotFoundError:
            # If the file is not found, create a new workbook with headers
            self.workbook = Workbook()
            self.sheet = self.workbook.active
            self.sheet.append(["Room Number", "Guest Name", "Check-in Date", "Check-out Date"])
            # Save the new workbook to the file
            self.workbook.save('hotel_data.xlsx')

    def check_availability(self):
        # Check if the number of occupied rooms is less than the total number of rooms
        return len(list(self.sheet.iter_rows(min_row=2, max_col=1, max_row=self.total_rooms))) < self.total_rooms

    def assign_room(self, guest_name, checkout_date):
        if self.check_availability():
            # Calculate the room number based on the existing rows in the sheet
            room_number = self.sheet.max_row - 10 if self.sheet.max_row > 8 else 1  # Subtracting 8 or start from 1 if the sheet is empty
            checkin_date = datetime.now().strftime("%Y-%m-%d")

            try:
                # Validate the checkout date format
                checkout_date = datetime.strptime(checkout_date, "%Y-%m-%d")
            except ValueError:
                print("Error: Invalid date format. Please use YYYY-MM-DD.")
                return

            if checkout_date < datetime.now():
                print("Error: Check-out date cannot be before the current date.")
                return

            # Append guest information to the sheet and save the workbook
            self.sheet.append([room_number, guest_name, checkin_date, checkout_date.strftime("%Y-%m-%d")])
            self.workbook.save('hotel_data.xlsx')
            print(f"Room assigned successfully. Room Number: {room_number}")
        else:
            print("Error: Sorry, the hotel is full. Cannot assign a room.")

    def display_guests(self):
        print("\nCurrent Guests:")
        for row in self.sheet.iter_rows(min_row=2, max_col=4):
            # Extract information from the row
            room_number, guest_name, checkin_date, checkout_date = [cell.value for cell in row]

            # Use .get() to handle potential None values more effectively
            room_number_str = str(room_number) if room_number is not None else ""
            guest_name_str = str(guest_name) if guest_name is not None else ""
            checkin_date_str = str(checkin_date) if checkin_date is not None else ""
            checkout_date_str = str(checkout_date) if checkout_date is not None else ""

            # Check if all values are None, and if so, continue to the next iteration
            if not any([room_number, guest_name, checkin_date, checkout_date]):
                continue

            # Display guest information
            print(f"Room {room_number_str}: {guest_name_str} ({checkin_date_str} to {checkout_date_str})")

# Example usage:
hotel_system = HotelManagementSystem()

while True:
    print("\n1. Check Room Availability")
    print("2. Assign Room")
    print("3. Display Current Guests")
    print("4. Exit")
    choice = input("Enter your choice (1/2/3/4): ")

    if choice == "1":
        # Check and display room availability
        if hotel_system.check_availability():
            print("Rooms are available.")
        else:
            print("Sorry, the hotel is full.")
    elif choice == "2":
        # Assign a room to a guest
        if hotel_system.check_availability():
            guest_name = input("Enter guest name: ")
            checkout_date = input("Enter check-out date (YYYY-MM-DD): ")
            hotel_system.assign_room(guest_name, checkout_date)
        else:
            print("Sorry, the hotel is full. Cannot assign a room.")
    elif choice == "3":
        # Display information about current guests
        hotel_system.display_guests()
    elif choice == "4":
        # Exit the program
        break
    else:
        print("Invalid choice. Please enter a valid option.")
